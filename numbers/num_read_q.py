# -*- coding: utf-8 -*-

import os
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from num_solver import num_solver
from num_template import num_template
		
class num_read_q:
	""" 数独問題読み込みクラス """
	version = '1.1.0'
	def __init__(self, _path, _temp):
		""" 初期化 """
		# テンプレートの読み込み
		for  t, te in enumerate(_temp):
			if t==0:
				self.te = num_template(te)
			else:
				self.te.read_xlsx(te)
		
		# 問題の読み込み
		book = openpyxl.load_workbook(_path)
		
		# 結果を保持(書き出し用)
		self.results = []
		
		for sheet in book.worksheets:
			# シート名をキーにする
			name = sheet.title
			print(name)
			
			# テンプレートの選択
			te = self.te.get( sheet['A1'].value )
			
			# テンプレートがないときは飛ばす
			if te==None:
				self.results.append([name, sheet['A1'].value])
				continue
			
			# 問題クラスの初期化
			nq = num_solver(te)
			
			# テンプレートに従って問題の読み込み
			for idx in te['template']:
				num = sheet[idx].value
				# 空欄のとき
				if num==None:
					num=-1
				nq.set(idx, _num=num)
			
			# 問題クラスの構築
			nq.configure(te)
			
			# 問題を解く
			nq.solve()
			
			# 結果を保持
			self.results.append([name, nq.result])
			
			# 解答を書き出す
			for idx, n in nq.nb.items():
				if n.is_ok():
					sheet[idx].value= n.num
					if n.init==False:
						sheet[idx].font = Font(bold=True, color='FF0000' )
				else:
					val ='('
					for c in  n.cand:
						val+='{},'.format(c)
					val +=')'
					sheet[idx].value= val
					sheet[idx].fill = PatternFill(patternType='solid', fgColor='FFFF00')
		
		# 結果を保存
		book.create_sheet(title='summary')
		summary = book.worksheets[-1]
		book.active = summary
		
		summary['B1'].value = 'version'
		summary['A2'].value = '全体'
		summary['B2'].value = num_read_q.version
		summary['A3'].value = 'solver'
		summary['B3'].value = num_solver.version
		
		for  t, te in enumerate(_temp):
			tbk = openpyxl.load_workbook(te)
			ws = tbk['version']
			cell_A = 'A{}'.format(t+4)
			cell_B = 'B{}'.format(t+4)
			cell_C = 'C{}'.format(t+4)
			
			summary[cell_A].value = 'template'
			summary[cell_B].value = ws['A2'].value
			summary[cell_C].value = os.path.basename(te)
		
		for r, ret in enumerate(self.results):
			cell_A = 'A{}'.format(r+5+len(_temp))
			cell_B = 'B{}'.format(r+5+len(_temp))
			
			summary[cell_A].value = ret[0]
			summary[cell_B].value = ret[1]
			
		# 別名で保存
		path2 = os.path.splitext(_path)
		book.save(path2[0] + '_solve' + path2[1])
			
if __name__ == "__main__":
	q_book = [
		#['./sample.xlsx', ['./template.xlsx']],
		['./sample2.xlsx', ['./template2.xlsx']],
		#['./問題/ナンプレ_20240306.xlsx', ['./template.xlsx', './問題/template_ナンプレ_20240306.xlsx']],
		#['./問題/ナンプレランド_20240210.xlsx', ['./template.xlsx', './問題/template_ナンプレランド_20240210.xlsx']],
		#['./問題/懸賞ナンプレ54_20250305.xlsx', ['./template.xlsx', './問題/template_懸賞ナンプレ54_20250305.xlsx']],
		#['./問題/ナンプレパークファミリー_20250307.xlsx', ['./template.xlsx', './問題/template_ナンプレパークファミリー_20250307.xlsx']],
		#['./問題/ナンプレランド_20250410.xlsx', ['./template.xlsx', './問題/template_ナンプレランド_20250410.xlsx']],
		
	]
	for qb, te in q_book:
		print(qb)
		num_read_q(qb, te)
