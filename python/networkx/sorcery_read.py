# -*- coding: utf-8 -*-

import os
import openpyxl

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
from xlsxwriter.utility import xl_range

#
# エクセル読み込み(Edge)
#
def read_edge_sheet(edge_sheet):
	# 1行目はヘッダ
	edge = []
	for i in range(2, len(tuple(edge_sheet.rows))+1):
		# 始点ノード
		src = int(edge_sheet.cell(i,1).value)
		
		# 終点ノード
		dst = int(edge_sheet.cell(i,2).value)
		
		dic = {}
		# エッジラベル
		label = edge_sheet.cell(i,3).value
		if label!= None:
			dic['label'] = label
		
		# 選択肢なしフラグ
		flag = edge_sheet.cell(i,4).value
		if flag==1:
			dic['color'] = (1.0,0.0,0.0)

		edge.append((src, dst, dic))
		
	return edge

#
# エクセル読み込み(Node)
#
def read_node_sheet(node_sheet):
	# 1行目はヘッダ
	node = {}
	for i in range(2, len(tuple(node_sheet.rows))+1):
		# ノード
		n = int(node_sheet.cell(i,1).value)
		
		# 大名称
		map = node_sheet.cell(i,2).value
		
		node_stat = {
			# 位置
			'pos' : (
				int(node_sheet.cell(i,4).value),
				int(node_sheet.cell(i,5).value)),
		}
		
		# ノードラベル
		label = node_sheet.cell(i,3).value
		if label!=None:
			node_stat['label'] = label
	
		# スタート-エンドフラグ
		s_e = node_sheet.cell(i,6).value
		#if s_e!=None:
		node_stat['s_e'] = s_e
		
		# ステータス
		status = node_sheet.cell(i,7).value
		if status==-1:
			# 死亡(赤)
			node_stat['color'] = (1.0,0.0,0.0)
		elif status==4:
			# アイテム入手(緑)
			node_stat['color']  = (0.0,1.0,0.0)
		elif status==-3:
			# 呪い(紫)
			node_stat['color']  = (0.5,0.0,0.5)
		elif status==0:
			# リブラ(水色)
			node_stat['color']  = (0.0,1.0,1.0)
		elif status==1:
			# 戦闘(黄)
			node_stat['color']  = (1.0,1.0,0.0)
		elif status==-2:
			# 空腹、時間経過()
			node_stat['color']  = (1.0,0.5,0.375)
			
		# サイズ
		x = node_sheet.cell(i,8).value
		y = node_sheet.cell(i,9).value
		if x!=None and y!=None:
			node_stat['figsize'] = (x, y)
		
		# 出力に追加
		if map in node.keys():
			node[map]['node'].append( (n, node_stat) )
		else:
			node[map] = { 'node' : [ (n , node_stat) ]}
		
	return node

#
# エクセル読み込み
#
def read_xlsx(excelFileName):
	# エクセルファイルの読み込み
	book = openpyxl.load_workbook(excelFileName)

	# 全体マップ
	edge_all_sheet = book['NodeIDEdge']
	edge_all = read_edge_sheet(edge_all_sheet)
	
	node_all_sheet = book['NodeID']
	node_all = read_node_sheet(node_all_sheet)
	
	# 各マップ
	node_sheet = book['Node']
	node = read_node_sheet(node_sheet)
	
	edge_sheet = book['Edge']
	edge = read_edge_sheet(edge_sheet)
	
	# ファイルを閉じる
	book.close()
	
	# ノードIDの設定
	node_all['全体図']['oder'] = 0
	node_id_dic = { node_stat['label'] : n for n, node_stat in node_all['全体図']['node'] }
	for map in node.keys():
		node[map]['oder'] = node_id_dic[map]
	
	# figsizeを書き換え
	node_all['全体図']['figsize'] = (20, 20)
	node_figsize_dic = { node_stat['label'] : node_stat['figsize'] for n, node_stat in node_all['全体図']['node'] }
	for map in node.keys():
		node[map]['figsize'] = node_figsize_dic[map]
	
	return node, edge, node_all, edge_all

#
# 使用した術の読み込み
#
def read_using_spel(node_sheet):
	using_spel = {}
	for i in range(2, len(tuple(node_sheet.rows))+1):
		# 使用した術を読み込み
		spel = node_sheet.cell(i,10).value
		spel_effect = node_sheet.cell(i,11).value
		if spel!=None:
			if spel in  using_spel.keys():
				using_spel[spel] ['cnt'] += 1
			else:
				using_spel[spel] = {'cnt' : 1,
					'success' : 0,
					'fault' : 0,
					'no_item' : 0,
					'no_use' : 0,
					'no_spel' : 0,
					}
			
			if spel_effect!=None:
				if spel_effect==10:
					using_spel[spel]['success'] += 1
				elif spel_effect==11:
					using_spel[spel]['fault'] += 1
				elif spel_effect==12:
					using_spel[spel]['no_item'] += 1
				elif spel_effect==14:
					using_spel[spel]['no_use'] += 1
				elif spel_effect==13:
					using_spel[spel]['no_spel'] += 1
	
	return using_spel

#
# 術シートの読み込み
#
def read_spel_sheet(spel_sheet):
	# 全術の読み込み
	spel = []
	for i in range(2, len(tuple(spel_sheet.rows))+1):
		spel_name = spel_sheet.cell(i,1).value
		spel_cost = spel_sheet.cell(i,2).value
		spel_tool = spel_sheet.cell(i,3).value
		if spel_tool==None:
			spel_tool = ''
		
		spel.append([
			spel_name, spel_cost, spel_tool,
			])
		
	return spel

#
# 術の統計結果を出力
#
def write_spel_sheet(spel, using_spel, book, sheet):
	
	# ヘッダ
	sheet.write(0, 0, '術')
	sheet.write(0, 1, '体力')
	sheet.write(0, 2, '必要アイテム')
	sheet.write(0, 3, '回数')
	sheet.write(0, 4, '成功')
	sheet.write(0, 5, '失敗')
	sheet.write(0, 6, '道具なし')
	sheet.write(0, 7, '無効')
	
	# 術を出力
	for s, sp in enumerate(spel):
		for i in range(3):
			sheet.write(s+1, i, sp[i])
		
		if sp[0] in using_spel.keys():
			sheet.write(s+1, 3,
				int( using_spel[ sp[0] ]['cnt']) )
			sheet.write(s+1, 4,
				int( using_spel[ sp[0] ]['success']) )
			sheet.write(s+1, 5,
				int( using_spel[ sp[0] ]['fault']) )
			sheet.write(s+1, 6,
				int( using_spel[ sp[0] ]['no_item']) )
			sheet.write(s+1, 7,
				int( using_spel[ sp[0] ]['no_use']) )
			
	# 無い術を出力
	spel_list = list(zip(*spel))[0]
	cnt = len(spel)+1
	for k,v in using_spel.items():
		if not k in spel_list:
			sheet.write(cnt, 0, k )
			#sheet.write(cnt, 1, '' )
			sheet.write(cnt, 3, int( v['cnt']) )
			cnt += 1
			
	# 使用回数を合計
	spel_cnt = { 
		'success' : 0,
		'fault' : 0,
		'no_item' : 0,
		'no_use' : 0,
		'no_spel' : 0,
	}
	for sk in spel_cnt.keys():
		for v in using_spel.values():
			spel_cnt[sk] += v[sk]

	# グラフで出力
	chart = book.add_chart({'type':'column'})
	chart.add_series({
		'name':'回数',
		'categories':[sheet.name, 1, 0, len(spel)-1, 0 ],
		'values':[sheet.name, 1, 3, len(spel)-1, 3 ],
		})

	chart.set_title({'name':'術分布'})
	chart.set_x_axis({'num_font':{'size':8}})
	chart.set_y_axis({'major_gridlines':{'visible':True},})
	chart.set_size({'width': 720, 'height': 576})
	sheet.insert_chart(xl_rowcol_to_cell(1, 9), chart, {  })

	chart2 = book.add_chart({'type':'column', 'subtype':'stacked'})
	chart2.add_series({
		'name':'成功',
		'categories':[sheet.name, 1, 0, len(spel)-1, 0 ],
		'values':[sheet.name, 1, 4, len(spel)-1, 4 ],
		})
	chart2.add_series({
		'name':'失敗',
		'categories':[sheet.name, 1, 0, len(spel)-1, 0 ],
		'values':[sheet.name, 1, 5, len(spel)-1, 5 ],
		})
	chart2.add_series({
		'name':'道具なし',
		'categories':[sheet.name, 1, 0, len(spel)-1, 0 ],
		'values':[sheet.name, 1, 6, len(spel)-1, 6 ],
		})
	chart2.add_series({
		'name':'無効',
		'categories':[sheet.name, 1, 0, len(spel)-1, 0 ],
		'values':[sheet.name, 1, 7, len(spel)-1, 7 ],
		})

	chart2.set_title({'name':'術分布'})
	chart2.set_x_axis({'num_font':{'size':7}})
	chart2.set_y_axis({'major_gridlines':{'visible':True},})
	chart2.set_size({'width': 720, 'height': 576})
	sheet.insert_chart(xl_rowcol_to_cell(35, 9), chart2, {  })
	
	# 円グラフ作成
	sheet.write(68, 9, '成功' )
	sheet.write(69, 9, '失敗' )
	sheet.write(70, 9, '道具なし' )
	sheet.write(71, 9, '無効' )
	sheet.write(72, 9, '術無し' )
	sheet.write(68, 10,  spel_cnt['success'])
	sheet.write(69, 10,  spel_cnt['fault'])
	sheet.write(70, 10,  spel_cnt['no_item'])
	sheet.write(71, 10,  spel_cnt['no_use'])
	sheet.write(72, 10,  spel_cnt['no_spel'])
	
	chart_pie = book.add_chart({'type':'pie'})
	chart_pie.add_series({
		'name':'術割合',
		'categories':[sheet.name, 68, 9, 72, 9 ],
		'values':[sheet.name, 68, 10, 72, 10 ],
		})
	chart_pie.set_size({'width': 720, 'height': 576})
	sheet.insert_chart(xl_rowcol_to_cell(74, 9), chart_pie, {  })

#
# 術の統計情報
#
def spel_stat(excelFileName):
	# エクセルファイルの読み込み
	book = openpyxl.load_workbook(excelFileName)
	
	# 術シートの読み込み
	spel_sheet = book['術']
	spel = read_spel_sheet(spel_sheet)
	
	# Nodeシートから使用した術の読み込み
	node_sheet = book['Node']
	using_spel = read_using_spel(node_sheet)
	
	# ファイルを閉じる
	book.close()
	
	return spel, using_spel
	
	
#
# メイン処理
#
def main(excelFileName):
	# エクセルファイルの読み込み
	node, edge, _, _ = read_xlsx(excelFileName)

def main2():
	# 出力ファイルの作成
	book = xlsxwriter.Workbook('Sorcery_術分析.xlsx')
	
	# 術の統計情報
	exec_list = [
		['Sorcery01.xlsx', 'vol1'],
		['Sorcery02.xlsx', 'vol2'],
		['Sorcery03.xlsx', 'vol3'],
		#['Sorcery04.xlsx', 'vol4'],
	]
	
	for f, t in exec_list:
		# 術の結果を読み込む
		spel, using_spel = spel_stat(f)
		
		# 統計情報を書き込む
		sheet = book.add_worksheet( t )
		write_spel_sheet(spel, using_spel, book, sheet)
	
	# ファイルの保存
	book.close()
	
if __name__ == '__main__':
	
	#main('Sorcery01.xlsx')
	main2()
	
	